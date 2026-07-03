import os
import re
from datetime import datetime
from pathlib import Path
import tempfile
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from flask import Flask, request, render_template, send_file, redirect, url_for, flash, session

from functools import wraps
import psycopg2
from psycopg2.extras import RealDictCursor
from werkzeug.security import generate_password_hash, check_password_hash

import subprocess
import shutil
from werkzeug.utils import secure_filename
# ============================================================
# CONFIGURACAO
# ============================================================

COLUNAS_FINAIS = [
    "Data Contabilização",
    "UG Emitente",
    "UG 2",
    "Observação",
    "Processo",
    "Estorno",
    "Tipo Patrimonial",
    "Item Patrimonial",
    "Operação Patrimonial",
    "Classificação Complementar",
    "Valor do item",
]

COLUNAS_CLASSIFICACAO = [
    "Identificação do Exercício",
    "Domicílio bancário UG",
    "Fonte",
    "Transferência de Despesa",
    "(UG2) Unidade Gestora",
    "Vinculação de Pagamento",
    "Fonte Detalhada",
]

OBSERVACAO_PADRAO = "Resgate e Aplicação Financeira no Mês"
PROCESSO_PADRAO = "GPR"
ESTORNO_PADRAO = "0"
TIPO_PATRIMONIAL_PADRAO = "106"
ITEM_PATRIMONIAL_PADRAO = "1045"

TOLERANCIA_VALOR = 0.01


# ============================================================
# UTILITARIOS GERAIS
# ============================================================

def normalizar_nome(valor):
    texto = str(valor or "").strip().lower()
    texto = texto.replace("_", " ").replace("-", " ")
    texto = re.sub(r"\s+", " ", texto)
    acentos = {
        "á": "a", "à": "a", "ã": "a", "â": "a",
        "é": "e", "ê": "e",
        "í": "i",
        "ó": "o", "ô": "o", "õ": "o",
        "ú": "u",
        "ç": "c",
    }
    for a, b in acentos.items():
        texto = texto.replace(a, b)
    return texto


def achar_coluna(df, nomes_possiveis, fallback_index=None, obrigatoria=False):
    mapa = {normalizar_nome(col): col for col in df.columns}

    for nome in nomes_possiveis:
        chave = normalizar_nome(nome)
        if chave in mapa:
            return mapa[chave]

    for nome in nomes_possiveis:
        chave = normalizar_nome(nome)
        for chave_real, col_real in mapa.items():
            if chave in chave_real or chave_real in chave:
                return col_real

    if fallback_index is not None and fallback_index < len(df.columns):
        return df.columns[fallback_index]

    if obrigatoria:
        raise ValueError(f"Coluna obrigatória não encontrada: {nomes_possiveis}")

    return None


def valor_celula(row, coluna, padrao=""):
    if coluna is None:
        return padrao
    try:
        valor = row[coluna]
    except Exception:
        return padrao
    if pd.isna(valor):
        return padrao
    return str(valor).strip()


def texto_para_numero(valor):
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)) and not pd.isna(valor):
        return float(valor)

    texto = str(valor).strip()
    if texto == "" or texto.lower() == "nan":
        return 0.0

    texto = texto.replace("R$", "").replace(" ", "")

    # Formato brasileiro: 1.234,56
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")

    # Remove caracteres que nao sejam numero, ponto ou sinal
    texto = re.sub(r"[^0-9.\-]", "", texto)

    if texto in ("", "-", "."):
        return 0.0

    try:
        return float(texto)
    except ValueError:
        return 0.0


def arredondar2(valor):
    return round(float(valor or 0), 2)


def valores_iguais(a, b, tolerancia=TOLERANCIA_VALOR):
    return abs(arredondar2(a) - arredondar2(b)) <= tolerancia


def limpar_parte_classificacao(valor):
    texto = str(valor or "").strip()
    texto = texto.replace("-", ".")
    texto = re.sub(r"\s*\.\s*", ".", texto)
    texto = re.sub(r"\s+", " ", texto)
    texto = re.sub(r"^[.\s]+|[.\s]+$", "", texto)
    return texto


def ajustar_blocos_classificacao(classificacao):
    if not classificacao:
        return ""

    partes = str(classificacao).split(".")

    # Ultimo bloco: 100 -> 000100
    if len(partes) >= 1:
        ultimo = partes[-1]
        if re.fullmatch(r"\d+", ultimo):
            partes[-1] = ultimo.zfill(6)

    # Bloco fonte detalhada: exemplo 10009 -> 010009
    # Indices finais: ... fonte_detalhada, vinculacao, ultimo
    if len(partes) >= 4:
        idx = len(partes) - 3
        bloco = partes[idx]
        if re.fullmatch(r"\d+", bloco):
            partes[idx] = bloco.zfill(6)

    return ".".join(partes)


def montar_classificacao(row, df):
    col_classificacao_pronta = achar_coluna(
        df,
        ["Classificação Complementar", "Classificacao Complementar"],
        fallback_index=None,
        obrigatoria=False,
    )

    if col_classificacao_pronta is not None:
        valor = valor_celula(row, col_classificacao_pronta)
        if valor:
            return ajustar_blocos_classificacao(limpar_parte_classificacao(valor))

    partes = []
    for nome in COLUNAS_CLASSIFICACAO:
        col = achar_coluna(df, [nome], fallback_index=None, obrigatoria=False)
        valor = valor_celula(row, col)
        limpo = limpar_parte_classificacao(valor)
        if limpo:
            partes.append(limpo)

    return ajustar_blocos_classificacao(".".join(partes))


def formatar_data(valor):
    if valor is None or str(valor).strip() == "" or str(valor).lower() == "nan":
        return ""

    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y")

    texto = str(valor).strip()

    # ja esta em dd/mm/yyyy
    if re.fullmatch(r"\d{2}/\d{2}/\d{4}", texto):
        return texto

    # data com hora dd/mm/yyyy hh:mm:ss
    m = re.match(r"^(\d{2}/\d{2}/\d{4})", texto)
    if m:
        return m.group(1)

    # yyy-mm-dd
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", texto)
    if m:
        return f"{m.group(3)}/{m.group(2)}/{m.group(1)}"

    dt = pd.to_datetime(texto, errors="coerce", dayfirst=True)
    if pd.isna(dt):
        return texto
    return dt.strftime("%d/%m/%Y")


def chave_data(valor):
    data = formatar_data(valor)
    m = re.fullmatch(r"(\d{2})/(\d{2})/(\d{4})", data)
    if not m:
        return data
    return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"


def nome_aba_por_data(data):
    data_fmt = formatar_data(data)
    m = re.fullmatch(r"(\d{2})/(\d{2})/(\d{4})", data_fmt)
    if m:
        return str(int(m.group(1)))
    return "Dia Atual"


def abrir_csv(caminho):
    encodings = ["utf-8-sig", "utf-8", "latin1", "cp1252"]
    seps = [";", ",", "\t"]

    ultimo_erro = None

    for enc in encodings:
        for sep in seps:
            try:
                df = pd.read_csv(caminho, sep=sep, encoding=enc, dtype=str)
                if df.shape[1] > 1:
                    df.columns = [str(c).strip().replace("\ufeff", "") for c in df.columns]
                    return df.fillna("")
            except Exception as e:
                ultimo_erro = e

    raise ValueError(f"Não consegui abrir o CSV: {caminho}\nErro: {ultimo_erro}")


# ============================================================
# DIA ATUAL - 016132
# ============================================================


def processar_dia_atual_016132(df132):
    """
    Dia atual - 016132.

    Regra correta:
    Processa apenas as colunas J, K e L.
    A coluna Total não entra no dia atual, para não duplicar valor no final.
    """
    return processar_dia_atual_por_blocos_jkl_total(df132)


def processar_dia_atual_por_blocos_jkl_total(df132):
    """
    Lê apenas as colunas J/K/L do relatório 016132.
    Para cada coluna que tiver valores, cria um bloco separado.

    J = Entradas e Saídas NP
    K = Entradas de GD
    L = Saídas OBE

    A coluna Total é ignorada no dia atual.
    """
    col_data = achar_coluna(
        df132,
        ["Data Contabilização", "Data Contabilizacao", "Data"],
        0,
        True
    )

    col_ug = achar_coluna(
        df132,
        ["Unidade Gestora", "UG Emitente", "UG"],
        1,
        True
    )

    blocos = [
        {
            "indice": 9,   # Coluna J
            "nome": "Entradas e Saídas NP"
        },
        {
            "indice": 10,  # Coluna K
            "nome": "Entradas de GD"
        },
        {
            "indice": 11,  # Coluna L
            "nome": "Saídas OBE"
        },
    ]

    linhas_saida = []

    for bloco in blocos:
        idx = bloco["indice"]

        if idx >= len(df132.columns):
            continue

        col_valor = df132.columns[idx]

        linhas_do_bloco = []

        for _, row in df132.iterrows():
            primeira_coluna = str(row.iloc[0] if len(row) else "").strip().upper()

            if primeira_coluna == "TOTAL":
                continue

            valor = texto_para_numero(row[col_valor])

            if arredondar2(valor) == 0:
                continue

            operacao = "1917" if valor < 0 else "1916"
            valor_item = abs(arredondar2(valor))

            # Mantém a classificação complementar original do relatório.
            classificacao = montar_classificacao(row, df132)

            linhas_do_bloco.append({
                "data": formatar_data(row[col_data]),
                "ug": valor_celula(row, col_ug),
                "operacao": operacao,
                "classificacao": classificacao,
                "valor_item": valor_item,
            })

        if not linhas_do_bloco:
            continue

        primeira_linha_do_bloco = True

        for item in linhas_do_bloco:
            if primeira_linha_do_bloco:
                linhas_saida.append([
                    item["data"],
                    item["ug"],
                    "",
                    OBSERVACAO_PADRAO,
                    PROCESSO_PADRAO,
                    ESTORNO_PADRAO,
                    TIPO_PATRIMONIAL_PADRAO,
                    ITEM_PATRIMONIAL_PADRAO,
                    item["operacao"],
                    item["classificacao"],
                    item["valor_item"],
                ])

                primeira_linha_do_bloco = False

            else:
                linhas_saida.append([
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    TIPO_PATRIMONIAL_PADRAO,
                    ITEM_PATRIMONIAL_PADRAO,
                    item["operacao"],
                    item["classificacao"],
                    item["valor_item"],
                ])

    return linhas_saida

def obter_ug_emitente_016132(df132):
    """
    Busca a UG Emitente válida no relatório 016132.
    Ignora linha de TOTAL.
    """
    col_ug = achar_coluna(
        df132,
        ["Unidade Gestora", "UG Emitente", "UG"],
        1,
        True
    )

    for _, row in df132.iterrows():
        primeira_coluna = str(row.iloc[0] if len(row) else "").strip().upper()

        if primeira_coluna == "TOTAL":
            continue

        ug = valor_celula(row, col_ug)

        if ug and ug.strip().upper() != "TOTAL":
            return ug

    return ""


def processar_dia_atual_por_total(df132):
    col_data = achar_coluna(df132, ["Data Contabilização", "Data Contabilizacao", "Data"], 0, True)
    col_ug = achar_coluna(df132, ["Unidade Gestora", "UG Emitente", "UG"], 1, True)
    col_total = achar_coluna(df132, ["Total"], None, True)

    linhas_saida = []
    primeira = True

    for _, row in df132.iterrows():
        primeira_coluna = str(row.iloc[0] if len(row) else "").strip().upper()
        if primeira_coluna == "TOTAL":
            continue

        total = texto_para_numero(row[col_total])
        if arredondar2(total) == 0:
            continue

        operacao = "1917" if total < 0 else "1916"
        valor_item = abs(arredondar2(total))
        classificacao = montar_classificacao(row, df132)

        if primeira:
            linhas_saida.append([
                formatar_data(row[col_data]),
                valor_celula(row, col_ug),
                "",
                OBSERVACAO_PADRAO,
                PROCESSO_PADRAO,
                ESTORNO_PADRAO,
                TIPO_PATRIMONIAL_PADRAO,
                ITEM_PATRIMONIAL_PADRAO,
                operacao,
                classificacao,
                valor_item,
            ])
            primeira = False
        else:
            linhas_saida.append([
                "", "", "", "", "", "",
                TIPO_PATRIMONIAL_PADRAO,
                ITEM_PATRIMONIAL_PADRAO,
                operacao,
                classificacao,
                valor_item,
            ])

    return linhas_saida


def processar_dia_atual_por_blocos_jkl(df132):
    """
    Regra alternativa do documento/print:
    lê colunas J/K/L, elimina zero e gera blocos.
    """
    col_data = achar_coluna(df132, ["Data Contabilização", "Data Contabilizacao", "Data"], 0, True)
    col_ug = achar_coluna(df132, ["Unidade Gestora", "UG Emitente", "UG"], 1, True)

    blocos = [
        (9, "Entradas e Saídas NP"),
        (10, "Entradas de GD"),
        (11, "Saídas OBE"),
    ]

    linhas_saida = []
    primeira = True

    for _, row in df132.iterrows():
        primeira_coluna = str(row.iloc[0] if len(row) else "").strip().upper()
        if primeira_coluna == "TOTAL":
            continue

        for idx, nome_padrao in blocos:
            if idx >= len(df132.columns):
                continue

            col_valor = df132.columns[idx]
            valor = texto_para_numero(row[col_valor])
            if arredondar2(valor) == 0:
                continue

            nome_bloco = str(col_valor).strip() or nome_padrao
            operacao = "1917" if valor < 0 else "1916"
            valor_item = abs(arredondar2(valor))

            # Quando for regra de bloco, a coluna J do template identifica o bloco.
            classificacao_ou_bloco = nome_bloco

            if primeira:
                linhas_saida.append([
                    formatar_data(row[col_data]),
                    valor_celula(row, col_ug),
                    "",
                    OBSERVACAO_PADRAO,
                    PROCESSO_PADRAO,
                    ESTORNO_PADRAO,
                    TIPO_PATRIMONIAL_PADRAO,
                    ITEM_PATRIMONIAL_PADRAO,
                    operacao,
                    classificacao_ou_bloco,
                    valor_item,
                ])
                primeira = False
            else:
                linhas_saida.append([
                    "", "", "", "", "", "",
                    TIPO_PATRIMONIAL_PADRAO,
                    ITEM_PATRIMONIAL_PADRAO,
                    operacao,
                    classificacao_ou_bloco,
                    valor_item,
                ])

    return linhas_saida


# ============================================================
# DATA FUTURA - 016199 -> 016174 -> 016170
# ============================================================

def preparar_016199(df199):
    """
    Print:
    016199: data coluna A, UG2 coluna G, Documento coluna J, Entradas OBO coluna K.
    Somar Entradas OBO para datas, UG2 e documentos iguais.
    """
    col_data = achar_coluna(df199, ["Data Contabilização", "Data Contabilizacao", "Data"], 0, True)
    col_ug2 = achar_coluna(df199, ["UG 2", "UG2", "Unidade Gestora 2", "(UG2) Unidade Gestora"], 6, True)
    col_doc = achar_coluna(df199, ["Documento", "Documento da Ordem Bancária", "Documento da Ordem Bancaria"], 9, True)
    col_entrada_obo = achar_coluna(df199, ["Entradas OBO", "Entrada OBO"], 10, True)

    grupos = {}

    for _, row in df199.iterrows():
        primeira_coluna = str(row.iloc[0] if len(row) else "").strip().upper()

        if primeira_coluna == "TOTAL":
            continue

        data = formatar_data(row[col_data])
        ug2 = valor_celula(row, col_ug2)
        documento = valor_celula(row, col_doc)
        valor = texto_para_numero(row[col_entrada_obo])

        if str(data).strip().upper() == "TOTAL":
            continue

        if str(documento).strip().upper() == "TOTAL":
            continue

        if not documento or arredondar2(valor) == 0:
            continue

        classificacao = montar_classificacao(row, df199)

        chave = (chave_data(data), ug2, documento, classificacao)

        if chave not in grupos:
            grupos[chave] = {
                "data_original": data,
                "ug2": ug2,
                "documento": documento,
                "classificacao": classificacao,
                "valor_obo": 0.0,
                "row_ref": row,
            }

        grupos[chave]["valor_obo"] += valor

    return list(grupos.values())


def preparar_016174(df174):
    """
    016174: Documento coluna D, Programacao de Desembolso coluna C,
    Data de Contabilizacao coluna B e valor de Pagamento.
    """
    col_ug = achar_coluna(df174, ["Unidade Gestora", "UG", "UG Emitente", "UG 2"], 0, False)
    col_data = achar_coluna(df174, ["Data Contabilização", "Data Contabilizacao", "Data"], 1, True)
    col_pd = achar_coluna(df174, ["Programação de Desembolso", "Programacao de Desembolso", "PD"], 2, True)
    col_doc = achar_coluna(df174, ["Documento", "Documento da Ordem Bancária", "Ordem Bancária", "Ordem Bancaria"], 3, True)
    col_pagamento = achar_coluna(df174, ["Pagamento", "Valor Pagamento", "Valor do Pagamento", "Valor"], 4, False)

    mapa_documento = {}

    for _, row in df174.iterrows():
        primeira_coluna = str(row.iloc[0] if len(row) else "").strip().upper()

        if primeira_coluna == "TOTAL":
            continue   

        documento = valor_celula(row, col_doc)
        pd_num = valor_celula(row, col_pd)
        if not documento or not pd_num:
            continue

        item = {
            "ug": valor_celula(row, col_ug),
            "data_contabilizacao": formatar_data(row[col_data]),
            "pd": pd_num,
            "documento": documento,
            "pagamento": texto_para_numero(row[col_pagamento]) if col_pagamento is not None else None,
        }

        mapa_documento.setdefault(documento, []).append(item)

    return mapa_documento


def preparar_016170(df170):
    """
    016170: PD coluna C e Data da Programacao.
    A data da programação precisa ser uma data válida, não documento OBO.
    """
    col_ug = achar_coluna(df170, ["Unidade Gestora", "UG", "UG Emitente", "UG 2"], 0, False)
    col_pd = achar_coluna(df170, ["Programação de Desembolso", "Programacao de Desembolso", "PD"], 2, True)

    # Primeiro tenta achar pelo nome correto da coluna
    col_data_pd = achar_coluna(
        df170,
        [
            "Data da Programação",
            "Data da Programacao",
            "Data Programação",
            "Data Programacao",
            "Data de Programação",
            "Data de Programacao"
        ],
        None,
        False
    )

    # Se não achar pelo nome, procura uma coluna que tenha cara de data,
    # para evitar pegar Documento/OBO por engano.
    if col_data_pd is None:
        col_data_pd = descobrir_coluna_data_programacao_016170(df170, col_pd)

    col_valor = achar_coluna(df170, ["Valor", "Pagamento", "Valor Pagamento", "Valor do Pagamento"], 4, False)

    mapa_pd = {}

    for _, row in df170.iterrows():
        primeira_coluna = str(row.iloc[0] if len(row) else "").strip().upper()

        if primeira_coluna == "TOTAL":
            continue

        pd_num = valor_celula(row, col_pd)

        if not pd_num:
            continue

        data_pd = formatar_data(row[col_data_pd]) if col_data_pd is not None else ""

        item = {
            "ug": valor_celula(row, col_ug),
            "pd": pd_num,
            "data_pd": data_pd,
            "valor": texto_para_numero(row[col_valor]) if col_valor is not None else None,
        }

        mapa_pd[pd_num] = item

        if item["ug"]:
            mapa_pd[(item["ug"], pd_num)] = item

    return mapa_pd

def descobrir_coluna_data_programacao_016170(df170, col_pd):
    """
    Procura uma coluna com valores em formato de data.
    Evita selecionar Documento/OBO como data.
    """
    melhor_coluna = None
    maior_qtd_datas = 0

    for col in df170.columns:
        # Não usa a própria coluna da PD
        if col == col_pd:
            continue

        qtd_datas = 0

        for _, row in df170.head(20).iterrows():
            valor = str(row[col]).strip()

            if valor.upper() == "TOTAL":
                continue

            # dd/mm/yyyy
            if re.match(r"^\d{2}/\d{2}/\d{4}", valor):
                qtd_datas += 1
                continue

            # yyyy-mm-dd
            if re.match(r"^\d{4}-\d{2}-\d{2}", valor):
                qtd_datas += 1
                continue

        if qtd_datas > maior_qtd_datas:
            maior_qtd_datas = qtd_datas
            melhor_coluna = col

    if melhor_coluna is None:
        raise ValueError(
            "Não consegui identificar a coluna de Data da Programação no relatório 016170. "
            "Confira se o CSV possui essa coluna."
        )

    return melhor_coluna

def escolher_174_por_valor(lista_174, valor_199):
    if not lista_174:
        return None

    # Preferencia: documento cujo pagamento bate com Entrada OBO do 016199.
    for item in lista_174:
        if item["pagamento"] is not None and valores_iguais(item["pagamento"], valor_199):
            return item

    # Se nao houver coluna de pagamento ou nao bater, usa o primeiro documento encontrado.
    return lista_174[0]


def buscar_pd_170(mapa_pd, item174):
    ug = item174.get("ug", "")
    pd_num = item174.get("pd", "")

    if ug and (ug, pd_num) in mapa_pd:
        return mapa_pd[(ug, pd_num)]

    return mapa_pd.get(pd_num)


def processar_data_futura(df199, df174, df170, ug_emitente_132):
    dados199 = preparar_016199(df199)
    mapa174 = preparar_016174(df174)
    mapa170 = preparar_016170(df170)

    grupos_finais = {}
    avisos = []

    # Soma o total do 016199 por documento.
    # Isso serve apenas para comparar com o pagamento do 016174.
    total_por_documento = {}

    for item199 in dados199:
        documento = item199["documento"]
        total_por_documento[documento] = total_por_documento.get(documento, 0.0) + item199["valor_obo"]

    for item199 in dados199:
        documento = item199["documento"]
        valor_linha_199 = item199["valor_obo"]
        valor_total_documento_199 = total_por_documento.get(documento, valor_linha_199)

        lista174 = mapa174.get(documento, [])

        # Agora a busca no 016174 usa o total do documento,
        # e não o valor quebrado por classificação.
        item174 = escolher_174_por_valor(lista174, valor_total_documento_199)

        if not item174:
            avisos.append(f"Documento {documento} do 016199 não encontrado no 016174.")
            continue

        if item174["pagamento"] is not None and not valores_iguais(item174["pagamento"], valor_total_documento_199):
            avisos.append(
                f"Documento {documento}: total 016199 ({valor_total_documento_199:.2f}) diferente do pagamento 016174 ({item174['pagamento']:.2f}). Usando mesmo assim."
            )
            # Não usa continue aqui.
            # Mesmo com diferença de centavos/agregação, ainda tenta localizar a PD.

        item170 = buscar_pd_170(mapa170, item174)

        if not item170:
            avisos.append(f"PD {item174['pd']} do documento {documento} não encontrada no 016170.")
            continue

        if item170["valor"] is not None and not valores_iguais(item170["valor"], valor_total_documento_199):
            avisos.append(
                f"PD {item174['pd']}: valor 016170 ({item170['valor']:.2f}) diferente do total 016199 ({valor_total_documento_199:.2f}). Usando mesmo assim."
            )

        # Regra principal da data futura:
        # usa a Data da Programação da PD.
        data_corrigida = item170["data_pd"] or item174["data_contabilizacao"] or item199["data_original"]

        classificacao = item199["classificacao"]

        chave = (
            chave_data(data_corrigida),
            classificacao,
        )

        if chave not in grupos_finais:
            grupos_finais[chave] = {
                "data": data_corrigida,
                "ug_emitente": ug_emitente_132,
                "classificacao": classificacao,
                "valor": 0.0,
            }

        grupos_finais[chave]["valor"] += valor_linha_199

    linhas = []
    grupos_ordenados = sorted(
        grupos_finais.values(),
        key=lambda x: (chave_data(x["data"]), str(x["classificacao"]))
    )

    primeira_linha = True

    for grupo in grupos_ordenados:
        valor = arredondar2(grupo["valor"])

        if valor == 0:
            continue

        operacao = "1917" if valor < 0 else "1916"
        valor_item = abs(valor)

        if primeira_linha:
            linhas.append([
                formatar_data(grupo["data"]),
                grupo["ug_emitente"],
                "",
                OBSERVACAO_PADRAO,
                PROCESSO_PADRAO,
                ESTORNO_PADRAO,
                TIPO_PATRIMONIAL_PADRAO,
                ITEM_PATRIMONIAL_PADRAO,
                operacao,
                grupo["classificacao"],
                valor_item,
            ])
            primeira_linha = False
        else:
            linhas.append([
                "",
                "",
                "",
                "",
                "",
                "",
                TIPO_PATRIMONIAL_PADRAO,
                ITEM_PATRIMONIAL_PADRAO,
                operacao,
                grupo["classificacao"],
                valor_item,
            ])

    return linhas, avisos
# ============================================================
# GERACAO DO XLSX
# ============================================================

def escrever_aba(wb, nome_aba, linhas):
    if nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        wb.remove(ws)

    ws = wb.create_sheet(nome_aba)

    ws.append(COLUNAS_FINAIS)
    for linha in linhas:
        ws.append(linha)

    formatar_aba(ws)
    return ws


def formatar_aba(ws):
    fill_header = PatternFill("solid", fgColor="D9D9D9")
    font_header = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_row = ws.max_row
    max_col = ws.max_column

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center

    ws.freeze_panes = "A2"

    # Formatos
    for row in range(2, max_row + 1):
        ws.cell(row=row, column=1).number_format = "dd/mm/yyyy"
        ws.cell(row=row, column=11).number_format = '#,##0.00'

    # Larguras aproximadas do modelo
    widths = {
        1: 16,
        2: 12,
        3: 10,
        4: 34,
        5: 12,
        6: 10,
        7: 16,
        8: 16,
        9: 18,
        10: 48,
        11: 16,
    }

    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Altura do cabecalho
    ws.row_dimensions[1].height = 22


def gerar_xlsx(caminho_132, caminho_199, caminho_174, caminho_170, caminho_saida):
    df132 = abrir_csv(caminho_132)

    linhas_atual = processar_dia_atual_016132(df132)

    if not linhas_atual:
        raise ValueError("Nenhuma linha foi gerada para o dia atual. Confira o CSV 016132.")

    tem_data_futura = (
        caminho_199 and str(caminho_199).strip()
        and caminho_174 and str(caminho_174).strip()
        and caminho_170 and str(caminho_170).strip()
    )

    linhas_futura = []
    avisos = []
    nome_aba_futura = ""

    if tem_data_futura:
        df199 = abrir_csv(caminho_199)
        df174 = abrir_csv(caminho_174)
        df170 = abrir_csv(caminho_170)

        ug_emitente_132 = obter_ug_emitente_016132(df132)

        linhas_futura, avisos = processar_data_futura(
            df199,
            df174,
            df170,
            ug_emitente_132
        )

        if not linhas_futura:
            avisos.append(
                "Nenhuma linha foi gerada para a data futura. "
                "O arquivo será gerado apenas com a aba do dia atual."
            )
            tem_data_futura = False

    wb = Workbook()

    # remove aba padrão
    ws_default = wb.active
    wb.remove(ws_default)

    nome_aba_atual = nome_aba_por_data(linhas_atual[0][0])

    escrever_aba(wb, nome_aba_atual, linhas_atual)

    if tem_data_futura and linhas_futura:
        nome_aba_futura = nome_aba_por_data(linhas_futura[0][0])

        if nome_aba_futura == nome_aba_atual:
            nome_aba_futura = f"{nome_aba_futura} Futuro"

        escrever_aba(wb, nome_aba_futura, linhas_futura)

    caminho_saida = Path(caminho_saida)
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho_saida)

    return {
        "saida": str(caminho_saida),
        "aba_atual": nome_aba_atual,
        "linhas_atual": len(linhas_atual),
        "aba_futura": nome_aba_futura,
        "linhas_futura": len(linhas_futura),
        "avisos": avisos,
        "tem_data_futura": tem_data_futura,
    }


# ============================================================
# APLICACAO WEB - FLASK
# ============================================================

app = Flask(__name__)

DATABASE_URL = os.environ.get("DATABASE_URL")


def conectar_banco():
    if not DATABASE_URL:
        raise RuntimeError("DATABASE_URL não configurada.")

    return psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)


def inicializar_banco():
    conn = conectar_banco()

    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS usuarios (
                        id SERIAL PRIMARY KEY,
                        nome TEXT NOT NULL,
                        email TEXT UNIQUE NOT NULL,
                        senha_hash TEXT NOT NULL,
                        ativo BOOLEAN NOT NULL DEFAULT TRUE,
                        admin BOOLEAN NOT NULL DEFAULT FALSE,
                        criado_em TIMESTAMP NOT NULL DEFAULT NOW()
                    );
                """)

                cur.execute("""
                    CREATE TABLE IF NOT EXISTS logs_acesso (
                        id SERIAL PRIMARY KEY,
                        usuario_id INTEGER REFERENCES usuarios(id),
                        acao TEXT NOT NULL,
                        ip TEXT,
                        user_agent TEXT,
                        criado_em TIMESTAMP NOT NULL DEFAULT NOW()
                    );
                """)

                admin_nome = os.environ.get("ADMIN_NOME", "Administrador")
                admin_email = os.environ.get("ADMIN_EMAIL")
                admin_senha = os.environ.get("ADMIN_PASSWORD")

                if admin_email and admin_senha:
                    cur.execute(
                        "SELECT id FROM usuarios WHERE email = %s",
                        (admin_email,)
                    )
                    usuario_existente = cur.fetchone()

                    if not usuario_existente:
                        cur.execute("""
                            INSERT INTO usuarios (nome, email, senha_hash, admin)
                            VALUES (%s, %s, %s, TRUE)
                        """, (
                            admin_nome,
                            admin_email,
                            generate_password_hash(admin_senha)
                        ))

    finally:
        conn.close()


def registrar_log(acao):
    usuario_id = session.get("usuario_id")
    ip = request.headers.get("X-Forwarded-For", request.remote_addr)
    user_agent = request.headers.get("User-Agent", "")

    try:
        conn = conectar_banco()
        with conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO logs_acesso (usuario_id, acao, ip, user_agent)
                    VALUES (%s, %s, %s, %s)
                """, (usuario_id, acao, ip, user_agent))
        conn.close()
    except Exception:
        pass


def login_obrigatorio(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if not session.get("usuario_id"):
            return redirect(url_for("login"))
        return func(*args, **kwargs)
    return wrapper

def admin_obrigatorio(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if not session.get("usuario_id"):
            return redirect(url_for("login"))

        if not session.get("usuario_admin"):
            flash("Acesso restrito ao administrador.", "erro")
            return redirect(url_for("index"))

        return func(*args, **kwargs)
    return wrapper

app.secret_key = os.environ.get("SECRET_KEY", "troque-esta-chave-em-producao")

ALLOWED_EXTENSIONS = {"csv"}


def arquivo_csv_valido(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def salvar_upload(file_storage, pasta_temp, nome_padrao):
    if not file_storage or not file_storage.filename:
        return ""

    if not arquivo_csv_valido(file_storage.filename):
        raise ValueError(f"O arquivo {file_storage.filename} não é CSV.")

    filename = secure_filename(file_storage.filename) or nome_padrao
    caminho = Path(pasta_temp) / filename
    file_storage.save(caminho)
    return str(caminho)

def localizar_libreoffice_web():
    """
    Localiza o LibreOffice no Windows ou Linux.
    No VPS/Linux, normalmente o comando será 'libreoffice' ou 'soffice'.
    """
    caminhos_possiveis = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        "/usr/bin/libreoffice",
        "/usr/bin/soffice",
        "/snap/bin/libreoffice",
    ]

    for caminho in caminhos_possiveis:
        if os.path.exists(caminho):
            return caminho

    # Tenta achar no PATH do sistema
    for comando in ["libreoffice", "soffice"]:
        caminho = shutil.which(comando)
        if caminho:
            return caminho

    return None


def converter_xlsx_para_xls_web(caminho_xlsx, pasta_saida):
    libreoffice = localizar_libreoffice_web()

    if not libreoffice:
        raise RuntimeError(
            "LibreOffice não encontrado. "
            "Instale o LibreOffice no servidor ou verifique se ele está disponível no PATH."
        )

    comando = [
        libreoffice,
        "--headless",
        "--convert-to",
        "xls",
        "--outdir",
        str(pasta_saida),
        str(caminho_xlsx),
    ]

    resultado = subprocess.run(
        comando,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        shell=False,
    )

    if resultado.returncode != 0:
        raise RuntimeError(
            "Erro ao converter o arquivo XLSX para XLS.\n\n"
            f"Saída:\n{resultado.stdout}\n\n"
            f"Erro:\n{resultado.stderr}"
        )

    caminho_xls = Path(pasta_saida) / (Path(caminho_xlsx).stem + ".xls")

    if not caminho_xls.exists():
        raise RuntimeError(
            "A conversão terminou, mas o arquivo .xls não foi encontrado."
        )

    return caminho_xls

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        senha = request.form.get("senha", "")

        conn = conectar_banco()

        try:
            with conn:
                with conn.cursor() as cur:
                    cur.execute("""
                        SELECT id, nome, email, senha_hash, ativo, admin
                        FROM usuarios
                        WHERE email = %s
                    """, (email,))

                    usuario = cur.fetchone()

            if not usuario:
                flash("E-mail ou senha inválidos.", "erro")
                return redirect(url_for("login"))

            if not usuario["ativo"]:
                flash("Usuário inativo.", "erro")
                return redirect(url_for("login"))

            if not check_password_hash(usuario["senha_hash"], senha):
                flash("E-mail ou senha inválidos.", "erro")
                return redirect(url_for("login"))

            session["usuario_id"] = usuario["id"]
            session["usuario_nome"] = usuario["nome"]
            session["usuario_email"] = usuario["email"]
            session["usuario_admin"] = usuario["admin"]

            registrar_log("login")

            return redirect(url_for("index"))

        finally:
            conn.close()

    return render_template("login.html")

@app.route("/admin/usuarios", methods=["GET", "POST"])
@login_obrigatorio
@admin_obrigatorio
def admin_usuarios():
    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        email = request.form.get("email", "").strip().lower()
        senha = request.form.get("senha", "")
        admin = True if request.form.get("admin") == "on" else False

        if not nome or not email or not senha:
            flash("Preencha nome, e-mail e senha.", "erro")
            return redirect(url_for("admin_usuarios"))

        try:
            conn = conectar_banco()

            with conn:
                with conn.cursor() as cur:
                    cur.execute("""
                        INSERT INTO usuarios (nome, email, senha_hash, admin, ativo)
                        VALUES (%s, %s, %s, %s, TRUE)
                    """, (
                        nome,
                        email,
                        generate_password_hash(senha),
                        admin
                    ))

            conn.close()
            registrar_log(f"criou_usuario:{email}")
            flash("Usuário criado com sucesso.", "sucesso")

        except Exception as e:
            flash(f"Erro ao criar usuário: {e}", "erro")

        return redirect(url_for("admin_usuarios"))

    usuarios = []

    try:
        conn = conectar_banco()

        with conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT id, nome, email, ativo, admin, criado_em
                    FROM usuarios
                    ORDER BY criado_em DESC
                """)
                usuarios = cur.fetchall()

        conn.close()

    except Exception as e:
        flash(f"Erro ao listar usuários: {e}", "erro")

    return render_template("usuarios.html", usuarios=usuarios)


@app.route("/admin/usuarios/<int:usuario_id>/editar", methods=["POST"])
@login_obrigatorio
@admin_obrigatorio
def editar_usuario(usuario_id):
    nome = request.form.get("nome", "").strip()
    email = request.form.get("email", "").strip().lower()
    senha = request.form.get("senha", "").strip()
    admin = True if request.form.get("admin") == "on" else False
    ativo = True if request.form.get("ativo") == "on" else False

    if not nome or not email:
        flash("Nome e e-mail são obrigatórios.", "erro")
        return redirect(url_for("admin_usuarios"))

    try:
        conn = conectar_banco()

        with conn:
            with conn.cursor() as cur:
                if senha:
                    cur.execute("""
                        UPDATE usuarios
                        SET nome = %s,
                            email = %s,
                            senha_hash = %s,
                            admin = %s,
                            ativo = %s
                        WHERE id = %s
                    """, (
                        nome,
                        email,
                        generate_password_hash(senha),
                        admin,
                        ativo,
                        usuario_id
                    ))
                else:
                    cur.execute("""
                        UPDATE usuarios
                        SET nome = %s,
                            email = %s,
                            admin = %s,
                            ativo = %s
                        WHERE id = %s
                    """, (
                        nome,
                        email,
                        admin,
                        ativo,
                        usuario_id
                    ))

        conn.close()
        registrar_log(f"editou_usuario:{email}")
        flash("Usuário atualizado com sucesso.", "sucesso")

    except Exception as e:
        flash(f"Erro ao editar usuário: {e}", "erro")

    return redirect(url_for("admin_usuarios"))


@app.route("/admin/usuarios/<int:usuario_id>/excluir", methods=["POST"])
@login_obrigatorio
@admin_obrigatorio
def excluir_usuario(usuario_id):
    if usuario_id == session.get("usuario_id"):
        flash("Você não pode excluir o próprio usuário logado.", "erro")
        return redirect(url_for("admin_usuarios"))

    try:
        conn = conectar_banco()

        with conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM usuarios WHERE id = %s", (usuario_id,))

        conn.close()
        registrar_log(f"excluiu_usuario_id:{usuario_id}")
        flash("Usuário excluído com sucesso.", "sucesso")

    except Exception as e:
        flash(f"Erro ao excluir usuário: {e}", "erro")

    return redirect(url_for("admin_usuarios"))

@app.route("/logout")
@login_obrigatorio
def logout():
    registrar_log("logout")
    session.clear()
    return redirect(url_for("login"))


@app.route("/", methods=["GET"])
@login_obrigatorio
def index():
    registrar_log("acessou_pagina_inicial")
    return render_template("index.html")


@app.route("/gerar", methods=["POST"])
@login_obrigatorio
def gerar_web():
    try:
        with tempfile.TemporaryDirectory() as pasta_temp:
            arquivo_132 = salvar_upload(request.files.get("arquivo_132"), pasta_temp, "016132.csv")
            arquivo_199 = salvar_upload(request.files.get("arquivo_199"), pasta_temp, "016199.csv")
            arquivo_174 = salvar_upload(request.files.get("arquivo_174"), pasta_temp, "016174.csv")
            arquivo_170 = salvar_upload(request.files.get("arquivo_170"), pasta_temp, "016170.csv")

            if not arquivo_132:
                raise ValueError("O CSV 016132 é obrigatório.")

            futuros = [arquivo_199, arquivo_174, arquivo_170]
            algum_futuro = any(futuros)
            todos_futuros = all(futuros)

            if algum_futuro and not todos_futuros:
                raise ValueError(
                    "Para gerar data futura, envie os 3 arquivos: 016199, 016174 e 016170. "
                    "Se não tiver data futura, deixe os 3 campos em branco."
                )

            nome_saida = f"TEMPLATE_APLICACAO_RESGATE_GPR_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            caminho_saida = Path(pasta_temp) / nome_saida

            resultado = gerar_xlsx(
                arquivo_132,
                arquivo_199,
                arquivo_174,
                arquivo_170,
                caminho_saida,
            )

            with open(resultado["saida"], "rb") as f:
                arquivo_memoria = BytesIO(f.read())

            arquivo_memoria.seek(0)

            registrar_log("gerou_xlsx")

            return send_file(
                arquivo_memoria,
                as_attachment=True,
                download_name=nome_saida,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    except Exception as e:
        flash(str(e), "erro")
        return redirect(url_for("index"))
@app.route("/converter-xls", methods=["POST"])
@login_obrigatorio
def converter_xls_web():
    try:
        with tempfile.TemporaryDirectory() as pasta_temp:
            arquivo_xlsx = request.files.get("arquivo_xlsx")

            if not arquivo_xlsx or arquivo_xlsx.filename.strip() == "":
                raise ValueError("Selecione um arquivo .xlsx para converter.")

            if not arquivo_xlsx.filename.lower().endswith(".xlsx"):
                raise ValueError("Envie um arquivo .xlsx válido.")

            nome_seguro = secure_filename(arquivo_xlsx.filename)
            caminho_xlsx = Path(pasta_temp) / nome_seguro
            arquivo_xlsx.save(caminho_xlsx)

            caminho_xls = converter_xlsx_para_xls_web(caminho_xlsx, pasta_temp)

            with open(caminho_xls, "rb") as f:
                arquivo_memoria = BytesIO(f.read())

            arquivo_memoria.seek(0)

            registrar_log("converteu_xls")

            return send_file(
                arquivo_memoria,
                as_attachment=True,
                download_name=caminho_xls.name,
                mimetype="application/vnd.ms-excel",
            )

    except Exception as e:
        flash(str(e), "erro")
        return redirect(url_for("index"))


with app.app_context():
    if DATABASE_URL:
        try:
            inicializar_banco()
            print("Banco inicializado com sucesso.")
        except Exception as e:
            print(f"Erro ao inicializar banco: {e}")
    else:
        print("DATABASE_URL não configurada. Banco não inicializado neste ambiente.")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8000, debug=True)